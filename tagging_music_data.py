
from common import load_track, GENRES
from keras.models import load_model
from optparse import OptionParser
import numpy as np
import os
from os import listdir
from os.path import isfile, join
from openpyxl import Workbook # xlsx library

def predict_genre(model_path, songs_path, output_path):
    ''' nacteni pisnicek a ulozeni do exceloveho souboru zanr pisnicek'''

    model = load_model(model_path)

    wb = Workbook()
    sheet = wb.active

    songs = [f for f in listdir(songs_path) if isfile(join(songs_path, f))]
    for index, song in enumerate(songs):
        song_data = load_track(os.path.join(songs_path, song))[0]

        song_data = np.expand_dims(song_data, axis=3)
        song_data = np.expand_dims(song_data, axis=0)
        result = model.predict(song_data)
        selected_genre = GENRES[np.argmax(result)]
        sheet.cell(row=index + 1, column=1).value = songs[index]
        sheet.cell(row=index +1, column=2).value = "-"
        sheet.cell(row=index +1, column=3).value = selected_genre

    wb.save(output_path)

if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option('-s', '--songs_path', dest='songs_path',
            default=os.path.join(os.path.dirname(__file__),
                'data/test_model/'),
            help='path to the song', metavar='Songs_PATH')
    parser.add_option('-m', '--model_path', dest='model_path',
            default=os.path.join(os.path.dirname(__file__),
                'models/model.h5'),
            help='path to model', metavar='MODEL_PATH')
    parser.add_option('-o', '--output_path', dest='output_path',
                      default=os.path.join(os.path.dirname(__file__),
                                           'results/result_spec.xlsx'),
                      help='results', metavar='RESULTS_PATH')
    options, args = parser.parse_args()

    predict_genre(options.model_path, options.songs_path, options.output_path)